import os
import json
from openpyxl import load_workbook

WORKING_DIR = os.path.dirname(os.path.realpath(__file__))

VISOR_IN_FILE = os.path.join(WORKING_DIR, 'VisorTransData.xlsx')
VISOR_FOLDER = os.path.join(WORKING_DIR, 'new_visor')
VISOR_OUT_FILE = os.path.join(WORKING_DIR, 'new_visor', 'visorTransData.json')
VISOR_DATA_FILE = os.path.join(WORKING_DIR, 'new_visor', 'visorData.json')

VISOR_LICENSE = 'LICENSE.md'
VISOR_DATA_DATA_KEY = 'data'
VISOR_DATA_UPDATE_COMIT_KEY = 'updateComitHash'

def stringToJson(filename : str, outputFile : str) -> None:

  wb = load_workbook(filename, read_only = True)

  stringData = {}
  for s in wb:
    rows = s.iter_rows(min_col = 1, min_row = 2, max_col = 17, max_row = None)
    headers = []
    for header in s[1]:
      if header.value:
        headers.append(header.value)

    for row in rows:
      name = row[0].value

      if not name:
        continue

      data = {}

      for i, string in enumerate(row[1:]):
        if string.value:
          # I hate excel why did I do this to myself
          data[i] = string.value.replace("\r", "").replace("_x000D_", "").replace("\\n", "\n")

      if data:
        stringData[name] = data

  with open(outputFile, "w") as f:
    json.dump(stringData, f, indent=4)

def update_visor_data()-> None:

  visor_dir = os.listdir(VISOR_FOLDER)

  visor_dir.remove('visorTransData.json')
  visor_dir.remove('visorData.json')

  visor_dir.sort()

  with open(os.path.join(VISOR_DATA_FILE), mode='w') as visor_json:
    json.dump(
      {VISOR_DATA_DATA_KEY:visor_dir, VISOR_DATA_UPDATE_COMIT_KEY:""},
      visor_json, indent=2, ensure_ascii=False)

def main()-> None:
  update_visor_data()
  stringToJson(VISOR_IN_FILE, VISOR_OUT_FILE)

if __name__ == "__main__":
  main()